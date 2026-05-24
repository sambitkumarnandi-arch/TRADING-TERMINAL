import math
import io
import numpy as np
import pandas as pd
import pytz
from tvDatafeed import TvDatafeed, Interval
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

def format_price(val):
    try:
        val = float(val)
        if val == 0: return "0.00"
        if abs(val) < 1:
            formatted = f"{val:.18f}".rstrip('0')
            return formatted + '0' if formatted.endswith('.') else formatted
        else:
            formatted = f"{val:.8f}".rstrip('0')
            if formatted.endswith('.'): return formatted + "00"
            if len(formatted.split('.')[1]) == 1: return formatted + "0"
            return formatted
    except:
        return str(val)

def set_excel_precision(cell, val):
    if isinstance(val, (int, float)):
        if abs(val) < 1: cell.number_format = '0.00##############'
        else: cell.number_format = '0.00######'

def create_11trend_excel(df, targets_dict, stats_dict, symbol_name):
    output = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'{symbol_name}_MASTER_DASHBOARD'
    df_excel = df.copy()
    if pd.api.types.is_datetime64tz_dtype(df_excel['DATE']):
        df_excel['DATE'] = df_excel['DATE'].dt.tz_localize(None)
    df_excel['DATE'] = df_excel['DATE'].dt.strftime('%Y-%m-%d')
    df_excel = df_excel.fillna("")
    font_header = Font(name='Segoe UI', bold=True, color='FFFFFF', size=11)
    font_bold = Font(name='Segoe UI', bold=True)
    align_c = Alignment(horizontal='center', vertical='center', wrap_text=True)
    border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    def apply_style(cell, fill_hex, is_header=False):
        cell.fill = PatternFill(start_color=fill_hex, end_color=fill_hex, fill_type='solid')
        cell.alignment = align_c; cell.border = border_thin
        if is_header: cell.font = font_header
    for col_idx, header in enumerate(df_excel.columns.tolist(), 1):
        c = ws.cell(row=1, column=col_idx, value=header)
        apply_style(c, '1E1E1E', True)
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 16
    for r_idx, row in enumerate(df_excel.values.tolist(), 2):
        for c_idx, val in enumerate(row, 1):
            c = ws.cell(row=r_idx, column=c_idx, value=val)
            apply_style(c, 'F2F2F2' if r_idx % 2 == 0 else 'FFFFFF')
            set_excel_precision(c, val)
    headers_stats = ['BASE CMP', 'AVERAGE_LN', 'AVERAGE_SQ_LN', 'VARIANCE', 'DAILY_VOL_%', 'EXPECTED_RANGE']
    for c_idx, header in enumerate(headers_stats, 9):
        c = ws.cell(row=1, column=c_idx, value=header)
        apply_style(c, '60497A', True)
        ws.column_dimensions[openpyxl.utils.get_column_letter(c_idx)].width = 20
    for c_idx, val in enumerate([stats_dict['Close'], stats_dict['Avg_LN'], stats_dict['Avg_SQ_LN'], stats_dict['Variance'], stats_dict['Vol'], stats_dict['Exp_Range']], 9):
        c = ws.cell(row=2, column=c_idx, value=val)
        apply_style(c, 'FFF2CC'); c.font = font_bold; set_excel_precision(c, val)
    trend_headers = ['DEGREE', 'DEG_FACTOR', 'CALC_RANGE', 'BUY_TARGET', 'SELL_TARGET']
    for c_idx, header in enumerate(trend_headers, 9):
        c = ws.cell(row=4, column=c_idx, value=header)
        fill_col = 'B97A57' if c_idx < 12 else ('00B050' if c_idx == 12 else 'C00000')
        apply_style(c, fill_col, True)
    degrees = [3.75, 7.5, 11.25, 15, 18.75, 22.5, 30, 45, 60, 75, 86.25]
    for idx, deg in enumerate(degrees, 5):
        row_data = [f"{deg}°", targets_dict['FACTOR'][deg], targets_dict['RANGE'][deg], targets_dict['BUY'][deg], targets_dict['SELL'][deg]]
        for c_idx, val in enumerate(row_data, 9):
            c = ws.cell(row=idx, column=c_idx, value=val)
            apply_style(c, 'F2F2F2' if idx % 2 == 0 else 'FFFFFF')
            if c_idx == 12:
                c.fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
                c.font = Font(color='007A33', bold=True)
            elif c_idx == 13:
                c.fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
                c.font = Font(color='C00000', bold=True)
            set_excel_precision(c, val)
        if deg in [15, 45, 86.25]:
            c_deg = ws.cell(row=idx, column=9)
            c_deg.fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
            c_deg.font = Font(bold=True, color='FF9900')
    wb.save(output); output.seek(0)
    return output

def get_11trend_data(symbol, market_type):
    symbol = symbol.upper().strip()
    ex_map = {
        'FOREX': ['OANDA','FOREXCOM','FX_IDC','FX','CAPITALCOM'],
        'BINANCE_SPOT': ['BINANCE','COINBASE','KUCOIN'],
        'BINANCE_PERP': ['BINANCE','BYBIT','OKX'],
        'INDIA_NSE': ['NSE','BSE'], 'INDIA_BSE': ['BSE','NSE'],
        'US_STOCKS': ['NASDAQ','NYSE','AMEX'],
        'GLOBAL_UNIVERSAL': ['NASDAQ','NYSE','TVC','CAPITALCOM','OANDA','BINANCE','NSE','BSE']
    }
    exchanges_to_try = ex_map.get(market_type, ['OANDA','BINANCE'])
    try:
        tv = TvDatafeed()
        df, successful_exchange = None, None
        for ex in exchanges_to_try:
            temp_df = tv.get_hist(symbol=symbol, exchange=ex, interval=Interval.in_daily, n_bars=30)
            if temp_df is not None and not temp_df.empty:
                df = temp_df.dropna(subset=['open','high','low','close']); successful_exchange = ex; break
        if df is None or df.empty or len(df) < 10: return None, "Not enough data."
        df_10_days = df.tail(10).copy().reset_index()
        df_10_days = df_10_days.rename(columns={'datetime':'DATE','open':'OPEN','high':'HIGH','low':'LOW','close':'CLOSE'})
        ist_tz = pytz.timezone('Asia/Kolkata')
        df_10_days['DATE'] = pd.to_datetime(df_10_days['DATE'])
        if df_10_days['DATE'].dt.tz is None:
            df_10_days['DATE'] = df_10_days['DATE'].dt.tz_localize('UTC').dt.tz_convert(ist_tz)
        else:
            df_10_days['DATE'] = df_10_days['DATE'].dt.tz_convert(ist_tz)
        df_10_days['PREV_CLOSE'] = df_10_days['CLOSE'].shift(1)
        df_10_days['LOG_RETURN'] = np.log(df_10_days['CLOSE'] / df_10_days['PREV_CLOSE'])
        df_10_days.loc[0, 'LOG_RETURN'] = 0.0
        df_10_days['SQ_LOG_RETURN'] = df_10_days['LOG_RETURN'] ** 2
        avg_ln = df_10_days['LOG_RETURN'].mean()
        avg_sq_ln = df_10_days['SQ_LOG_RETURN'].mean()
        variance = avg_sq_ln - (avg_ln ** 2)
        daily_vol_pct = np.sqrt(variance) * 100
        last_close = float(df_10_days['CLOSE'].iloc[-1])
        expected_range = (daily_vol_pct * last_close) / 100
        mult = 1
        if last_close > 0 and last_close < 100:
            magnitude = math.floor(math.log10(last_close))
            mult = 10 ** (2 - magnitude)
        scaled_expected_range = expected_range * mult
        degrees = [3.75, 7.5, 11.25, 15, 18.75, 22.5, 30, 45, 60, 75, 86.25]
        buy_targets = {}; sell_targets = {}; factors = {}; ranges_dict = {}
        for deg in degrees:
            deg_factor = deg / 180.0
            scaled_calc_range = (np.sqrt(scaled_expected_range) + deg_factor) ** 2
            calc_range = scaled_calc_range / mult
            buy_targets[deg] = last_close + (deg_factor * calc_range)
            sell_targets[deg] = last_close - (deg_factor * calc_range)
            factors[deg] = deg_factor
            ranges_dict[deg] = calc_range
        export_df = df_10_days[['DATE','OPEN','HIGH','LOW','CLOSE','LOG_RETURN','SQ_LOG_RETURN']].copy()
        live_price = last_close
        market_status = "OFF (Closed)"
        status_color = "#ff3366"
        try:
            live_df = tv.get_hist(symbol=symbol, exchange=successful_exchange, interval=Interval.in_1_minute, n_bars=5)
            if live_df is not None and not live_df.empty:
                live_price = float(live_df['close'].iloc[-1])
                market_status = "ON (Live)"
                status_color = "#00ff7f"
        except Exception: pass
        return {
            'Symbol': symbol, 'Exchange': successful_exchange,
            'Date_Range': f"{export_df['DATE'].iloc[0].strftime('%d %b')} to {export_df['DATE'].iloc[-1].strftime('%d %b')}",
            'Close': last_close, 'Volatility': daily_vol_pct, 'Expected_Range': expected_range,
            'Degrees_List': degrees,
            'Targets': {'BUY': buy_targets, 'SELL': sell_targets, 'FACTOR': factors, 'RANGE': ranges_dict},
            'Stats': {'Close': last_close, 'Avg_LN': avg_ln, 'Avg_SQ_LN': avg_sq_ln, 'Variance': variance, 'Vol': daily_vol_pct, 'Exp_Range': expected_range},
            'Dataframe': export_df, 'Live_Price': live_price, 'Market_Status': market_status, 'Status_Color': status_color
        }, "Success"
    except Exception as e:
        return None, f"System Error: {str(e)}"
