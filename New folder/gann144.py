import io
import pandas as pd
import pytz
from datetime import timedelta
from tvDatafeed import TvDatafeed, Interval
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

def clean_and_float(series):
    return pd.to_numeric(series.astype(str).str.replace(',', '').str.strip(), errors='coerce')

def set_excel_precision(cell, val):
    if isinstance(val, (int, float)):
        if abs(val) < 1: cell.number_format = '0.00##############'
        else: cell.number_format = '0.00######'

def create_gann_excel(df, symbol_name, buy_status, sell_status, multiplier):
    output = io.BytesIO()
    wb = openpyxl.Workbook()
    ws_raw = wb.active; ws_raw.title = 'RAW_DATA'
    df_excel = df.copy()
    if pd.api.types.is_datetime64tz_dtype(df_excel['DATE']):
        df_excel['DATE'] = df_excel['DATE'].dt.tz_localize(None)
    df_excel['DATE'] = df_excel['DATE'].dt.strftime('%Y-%m-%d %H:%M')
    ws_raw.append(df_excel.columns.tolist())
    for row in df_excel.values.tolist(): ws_raw.append(row)
    border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    align_c = Alignment(horizontal='center', vertical='center')
    def style_cell(cell, bg_hex, font_color='000000', bold=False):
        cell.fill = PatternFill(start_color=bg_hex, end_color=bg_hex, fill_type='solid')
        cell.font = Font(color=font_color, bold=bold, name='Segoe UI')
        cell.alignment = align_c; cell.border = border_thin
    ws_dash = wb.create_sheet('DASHBOARD')
    ws_dash.column_dimensions['A'].width = 25; ws_dash.column_dimensions['B'].width = 45
    ws_dash.merge_cells('A1:B1')
    c1 = ws_dash['A1']; c1.value = f"GANN 144: {symbol_name}"
    style_cell(c1, '1E1E1E', 'FFFFFF', True)
    ws_dash['A2'] = 'BUY STATUS:'; style_cell(ws_dash['A2'], '00B050', 'FFFFFF', True)
    ws_dash['B2'] = buy_status; style_cell(ws_dash['B2'], 'E2EFDA', '000000', False)
    ws_dash['A3'] = 'SELL STATUS:'; style_cell(ws_dash['A3'], 'C00000', 'FFFFFF', True)
    ws_dash['B3'] = sell_status; style_cell(ws_dash['B3'], 'FCE4D6', '000000', False)
    ws_matrix = wb.create_sheet('MATRIX')
    for col in range(1, 14):
        ws_matrix.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20
    headers = ['TIME CYCLE','PRICE CYCLE','PC * DIFF','RESISTANCE 1','RESISTANCE 2','RESISTANCE 3','RESISTANCE 4','RESISTANCE 5','SUPPORT 1','SUPPORT 2','SUPPORT 3','SUPPORT 4','SUPPORT 5']
    for c_idx, val in enumerate(headers, 1):
        cell = ws_matrix.cell(row=1, column=c_idx, value=val)
        style_cell(cell, '1E1E1E', 'FFFFFF', True)
    max_high = df['HIGH'].max() * multiplier
    min_low = df['LOW'].min() * multiplier
    diff = max_high - min_low
    time_cycles = [36, 45, 54, 63, 72, 81, 90, 108, 126, 144]
    price_cycles = [0.25, 0.3125, 0.375, 0.4375, 0.5, 0.5625, 0.625, 0.75, 0.875, 1.0]
    for i, (tc, pc) in enumerate(zip(time_cycles, price_cycles), start=2):
        active_diff = diff * pc
        ws_matrix.cell(row=i, column=1, value=tc)
        ws_matrix.cell(row=i, column=2, value=pc)
        ws_matrix.cell(row=i, column=3, value=active_diff)
        for col in [1,2,3]: style_cell(ws_matrix.cell(row=i, column=col), 'F2F2F2')
        set_excel_precision(ws_matrix.cell(row=i, column=3), active_diff)
        h1 = max_high - active_diff
        res_vals = [h1, h1-diff, h1-2*diff, h1-3*diff, h1-4*diff]
        for j, val in enumerate(res_vals, start=4):
            cell = ws_matrix.cell(row=i, column=j, value=val)
            style_cell(cell, 'FCE4D6')
            set_excel_precision(cell, val)
        l1 = min_low + active_diff
        sup_vals = [l1, l1+diff, l1+2*diff, l1+3*diff, l1+4*diff]
        for j, val in enumerate(sup_vals, start=9):
            cell = ws_matrix.cell(row=i, column=j, value=val)
            style_cell(cell, 'E2EFDA')
            set_excel_precision(cell, val)
    wb.save(output); output.seek(0)
    return output

def get_gann_data(symbol, market_type, start_date, end_date, timeframe):
    symbol = symbol.upper().strip()
    ex_map = {
        'FOREX': ['OANDA','FOREXCOM','FX_IDC','FX','CAPITALCOM'],
        'BINANCE_SPOT': ['BINANCE','COINBASE','KUCOIN'],
        'BINANCE_PERP': ['BINANCE','BYBIT','OKX'],
        'INDIA_NSE': ['NSE','BSE'], 'INDIA_BSE': ['BSE','NSE'],
        'US_STOCKS': ['NASDAQ','NYSE','AMEX'],
        'GLOBAL_UNIVERSAL': ['NASDAQ','NYSE','TVC','CAPITALCOM','OANDA','BINANCE','NSE','BSE']
    }
    exchanges_to_try = ex_map.get(market_type, ['OANDA','BINANCE'])
    try:
        tv = TvDatafeed()
        tv_interval = Interval.in_1_hour if timeframe == '1-Hour' else Interval.in_daily
        raw_df = None
        for ex in exchanges_to_try:
            temp_df = tv.get_hist(symbol=symbol, exchange=ex, interval=tv_interval, n_bars=5000)
            if temp_df is not None and not temp_df.empty:
                raw_df = temp_df; break
        if raw_df is None or raw_df.empty: return pd.DataFrame()
        raw_df = raw_df.reset_index()
        ist_tz = pytz.timezone('Asia/Kolkata')
        raw_df['datetime'] = pd.to_datetime(raw_df['datetime'])
        if raw_df['datetime'].dt.tz is None:
            raw_df['datetime'] = raw_df['datetime'].dt.tz_localize('UTC').dt.tz_convert(ist_tz)
        else:
            raw_df['datetime'] = raw_df['datetime'].dt.tz_convert(ist_tz)
        raw_df['DATE'] = raw_df['datetime'].dt.date
        mask = (raw_df['DATE'] >= start_date) & (raw_df['DATE'] <= end_date)
        df_filtered = raw_df.loc[mask].copy()
        if df_filtered.empty: return pd.DataFrame()
        df = pd.DataFrame()
        df['DATE'] = df_filtered['datetime']
        df['OPEN'] = clean_and_float(df_filtered['open'])
        df['HIGH'] = clean_and_float(df_filtered['high'])
        df['LOW'] = clean_and_float(df_filtered['low'])
        df['CLOSE'] = clean_and_float(df_filtered['close'])
        return df.dropna(subset=['OPEN'])
    except Exception:
        return pd.DataFrame()

def analyze_gann(df, multiplier):
    if df.empty: return None, "", ""
    max_high = df['HIGH'].max() * multiplier
    min_low = df['LOW'].min() * multiplier
    diff = max_high - min_low
    last_close = df['CLOSE'].iloc[-1] * multiplier
    b36 = min_low + (diff * 0.25)
    b90 = min_low + (diff * 0.625)
    b108 = min_low + (diff * 0.75)
    s36 = max_high - (diff * 0.25)
    s90 = max_high - (diff * 0.625)
    s108 = max_high - (diff * 0.75)
    if last_close >= b108: buy_status = "Crossed 108 in front. Wait for next iteration 36 Time Cycle to pass."
    elif b90 <= last_close < b108: buy_status = "Price crossed 108 moving backwards. Target moving towards 36 Cycle."
    elif b36 <= last_close < b90: buy_status = "Crossed Time Cycle 36. Target is 90 Cycle (Book Profit)."
    else: buy_status = "Awaiting Setup. Price is below 36 Cycle."
    if last_close <= s108: sell_status = "Crossed 108 downside. Wait for next 36 Time Cycle."
    elif s108 < last_close <= s90: sell_status = "Price crossed 108 upside. Target towards 36 Cycle."
    elif s90 < last_close <= s36: sell_status = "Crossed Time Cycle 36 downwards. Target is 90 Cycle."
    else: sell_status = "Awaiting Setup. Price is above 36 Cycle."
    return last_close, buy_status, sell_status
