import math
import io
import base64
import pytz
from datetime import timedelta, timezone, datetime
from tvDatafeed import TvDatafeed, Interval
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

def format_price(val):
    try:
        val = float(val)
        if val == 0: return "0.00"
        if abs(val) < 1:
            formatted = f"{val:.18f}".rstrip('0')
            return formatted + '0' if formatted.endswith('.') else formatted
        else:
            formatted = f"{val:.8f}".rstrip('0')
            if formatted.endswith('.'): return formatted + "00"
            if len(formatted.split('.')[1]) == 1: return formatted + "0"
            return formatted
    except:
        return str(val)

def set_excel_precision(cell, val):
    if isinstance(val, (int, float)):
        if abs(val) < 1: cell.number_format = '0.00##############'
        else: cell.number_format = '0.00######'

def create_midpoint_excel(data):
    output = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{data['Symbol']}_MIDPOINT"
    font_title = Font(name='Segoe UI', size=16, bold=True, color='FFFFFF')
    font_bold = Font(name='Segoe UI', bold=True)
    align_c = Alignment(horizontal='center', vertical='center')
    border_all = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    fill_dark = PatternFill(start_color='1E1E1E', end_color='1E1E1E', fill_type='solid')
    fill_buy = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
    fill_sell = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
    fill_yellow = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    ws.merge_cells('A1:D1')
    c1 = ws['A1']
    c1.value = f"MID-POINT TRADING ALGORITHM: {data['Symbol']} ({data['Exchange']})"
    c1.font, c1.fill, c1.alignment = font_title, fill_dark, align_c
    ws.row_dimensions[1].height = 25
    info = [("Base 1H Candle (IST)", data['Time']), ("High", data['High']), ("Low", data['Low']),
            ("Range", data['Range']), ("MID POINT", data['Mid']), ("STOP LOSS", data['Stop_Loss'])]
    for i, (label, val) in enumerate(info, start=3):
        ws[f'A{i}'] = label; ws[f'B{i}'] = val
        ws[f'A{i}'].font = font_bold; ws[f'B{i}'].font = font_bold
        ws[f'A{i}'].fill = fill_yellow
        ws[f'A{i}'].border = border_all; ws[f'B{i}'].border = border_all
        set_excel_precision(ws[f'B{i}'], val)
    ws['A11'] = "BUY ZONE"; ws.merge_cells('A11:B11')
    ws['A11'].fill = PatternFill(start_color='00B050', end_color='00B050', fill_type='solid')
    ws['A11'].font = Font(color='FFFFFF', bold=True); ws['A11'].alignment = align_c
    ws['C11'] = "SELL ZONE"; ws.merge_cells('C11:D11')
    ws['C11'].fill = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
    ws['C11'].font = Font(color='FFFFFF', bold=True); ws['C11'].alignment = align_c
    buy_data = [("ENTRY", data['Buy_Entry']), ("TARGET 1", data['Buy_T1']), ("MID TGT", data['Buy_between T1 & T2']), ("TARGET 2", data['Buy_T2']), ("TARGET 3", data['Buy_T3'])]
    sell_data = [("ENTRY", data['Sell_Entry']), ("TARGET 1", data['Sell_T1']), ("MID TGT", data['Sell_between T1 & T2']), ("TARGET 2", data['Sell_T2']), ("TARGET 3", data['Sell_T3'])]
    for i, ((bl, bv), (sl, sv)) in enumerate(zip(buy_data, sell_data), start=12):
        ws[f'A{i}'] = bl; ws[f'B{i}'] = bv; ws[f'C{i}'] = sl; ws[f'D{i}'] = sv
        for col in ['A','B','C','D']:
            cell = ws[f'{col}{i}']; cell.border = border_all; cell.font = font_bold
            if col in ['A','B']: cell.fill = fill_buy
            else: cell.fill = fill_sell
            if col in ['B','D']: set_excel_precision(cell, cell.value)
    ws.column_dimensions['A'].width = 22; ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 22; ws.column_dimensions['D'].width = 18
    wb.save(output); output.seek(0)
    return output

def get_midpoint_data(symbol, market_type):
    symbol = symbol.upper().strip()
    ex_map = {
        'FOREX': ['OANDA','FOREXCOM','FX_IDC','FX','CAPITALCOM','EIGHTCAP'],
        'BINANCE_SPOT': ['BINANCE','COINBASE','KUCOIN','CRYPTO','HYPERLIQUID'],
        'BINANCE_PERP': ['BINANCE','BYBIT','OKX','HYPERLIQUID'],
        'INDIA_NSE': ['NSE','BSE'],
        'INDIA_BSE': ['BSE','NSE'],
        'US_STOCKS': ['NASDAQ','NYSE','AMEX','BATS'],
        'GLOBAL_UNIVERSAL': ['NASDAQ','NYSE','TVC','CAPITALCOM','LSE','TSX','ASX','HKEX','TSE','XETR','EURONEXT','OANDA','BINANCE','NSE','BSE']
    }
    exchanges_to_try = ex_map.get(market_type, ['OANDA','BINANCE','HYPERLIQUID'])
    try:
        tv = TvDatafeed()
        df, successful_exchange = None, None
        for ex in exchanges_to_try:
            temp_df = tv.get_hist(symbol=symbol, exchange=ex, interval=Interval.in_1_hour, n_bars=2)
            if temp_df is not None and not temp_df.empty:
                df = temp_df; successful_exchange = ex; break
        if df is None or df.empty: return None, f"'{symbol}' not found."
        last_closed_candle = df.iloc[-2] if len(df) >= 2 else df.iloc[-1]
        open_price = float(last_closed_candle['open'])
        high_price = float(last_closed_candle['high'])
        low_price = float(last_closed_candle['low'])
        close_price = float(last_closed_candle['close'])
        mid_point = (high_price + low_price) / 2
        price_range = high_price - low_price
        half_range = high_price - mid_point
        mult = 1
        if mid_point > 0 and mid_point < 100:
            magnitude = math.floor(math.log10(mid_point))
            mult = 10 ** (2 - magnitude)
        scaled_half_range = half_range * mult
        offset_360 = ((math.sqrt(scaled_half_range) + 2)**2) / mult if scaled_half_range > 0 else 0
        uptrend_term = mid_point + offset_360
        downtrend_term = mid_point - offset_360
        buy_entry = mid_point + (price_range * 0.236)
        sell_entry = mid_point - (price_range * 0.236)
        return {
            'Symbol': symbol, 'Exchange': successful_exchange,
            'Time': datetime.now(timezone(timedelta(hours=5, minutes=30))).strftime('%d %b %Y, %I:%M %p IST'),
            'Open': open_price, 'High': high_price, 'Low': low_price, 'Close': close_price,
            'Mid': mid_point, 'Range': price_range,
            'Buy_Entry': buy_entry, 'Buy_T1': (high_price + mid_point)/2,
            'Buy_T2': (high_price + uptrend_term)/2,
            'Buy_T3': ((high_price + uptrend_term)/2 + uptrend_term)/2,
            'Buy_between T1 & T2': ((high_price + mid_point)/2 + (high_price + uptrend_term)/2)/2,
            'Sell_Entry': sell_entry, 'Sell_T1': (low_price + mid_point)/2,
            'Sell_T2': (low_price + downtrend_term)/2,
            'Sell_T3': ((low_price + downtrend_term)/2 + downtrend_term)/2,
            'Sell_between T1 & T2': ((low_price + mid_point)/2 + (low_price + downtrend_term)/2)/2,
            'Stop_Loss': mid_point, 'Up_Term': uptrend_term, 'Down_Term': downtrend_term
        }, "Success"
    except Exception as e:
        return None, f"System Error: {str(e)}"
